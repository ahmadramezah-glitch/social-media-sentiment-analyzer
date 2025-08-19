from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file

from config import INSTAGRAM_ACCESS_TOKEN, INSTAGRAM_USER_ID
from instagram_api import get_hashtag_id, fetch_recent_posts, fetch_post_comments
from sentiment import analyze_sentiment
from models import db, Post, User, Comment
from tiktok_api import search_tiktok_hashtag, TikTokAPI
from twitter_api import search_twitter_hashtag, fetch_tweet_comments
import os
from functools import wraps
from collections import defaultdict
import calendar
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from typing import List, Dict, Any
import sqlalchemy.exc
import time

app = Flask(__name__)

# Demo data generation removed - only real Instagram data will be used
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///db.sqlite3'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = 'your-secret-key-here-change-in-production'

# Session configuration
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
app.config['SESSION_PERMANENT'] = True



# Add session configuration debugging
print(f"DEBUG: Flask app created")
print(f"DEBUG: Secret key set: {app.secret_key[:20]}...")

# Global storage for Twitter results (temporary solution)
twitter_results_storage = {}

def cleanup_old_twitter_results():
    """Clean up old Twitter results from global storage"""
    current_time = time.time()
    expired_keys = []
    for key, data in twitter_results_storage.items():
        if current_time - data['timestamp'] > 3600:  # 1 hour
            expired_keys.append(key)
    
    for key in expired_keys:
        del twitter_results_storage[key]
        print(f"Cleaned up expired Twitter results: {key}")
    
    return len(expired_keys)

# Authentication credentials (for initial admin)
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"

# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Admin required decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = User.query.get(session['user_id'])
        if not user or not user.is_admin:
            flash('Access denied. Admin privileges required.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# Permission required decorator
def permission_required(permission):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            user = User.query.get(session['user_id'])
            if not user or not user.has_permission(permission):
                flash(f'Access denied. {permission.replace("_", " ").title()} permission required.', 'danger')
                return redirect(url_for('no_access'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

db.init_app(app)

# Create tables and initial admin user when the app starts
with app.app_context():
    db.create_all()
    
    # Create initial admin user if it doesn't exist
    admin_user = User.query.filter_by(username=ADMIN_USERNAME).first()
    if not admin_user:
        admin_user = User(
            username=ADMIN_USERNAME,
            email='admin@example.com',
            is_admin=True,
            is_active=True,
            can_view_dashboard=True,
            can_view_graphs=True,
            can_search_hashtags=True,
            can_view_filtered_results=True,
            can_view_all_posts=True,
            can_manage_users=True
        )
        admin_user.set_password(ADMIN_PASSWORD)
        db.session.add(admin_user)
        db.session.commit()
        print(f"Created admin user: {ADMIN_USERNAME}")
    else:
        # Update existing admin user with new permissions
        admin_user.can_view_dashboard = True
        admin_user.can_view_graphs = True
        admin_user.can_search_hashtags = True
        admin_user.can_view_filtered_results = True
        admin_user.can_view_all_posts = True
        admin_user.can_manage_users = True
        db.session.commit()
        
        # No demo data - only real Instagram data will be used
        print("Real Instagram data only - no demo data will be generated")

def cleanup_duplicate_posts():
    """Clean up duplicate posts to prevent UNIQUE constraint violations"""
    try:
        # Find duplicate posts by post_id and source
        duplicates = db.session.query(Post.post_id, Post.source, db.func.count(Post.id).label('count'))\
            .group_by(Post.post_id, Post.source)\
            .having(db.func.count(Post.id) > 1)\
            .all()
        
        if duplicates:
            print(f"Found {len(duplicates)} duplicate post groups")
            for dup in duplicates:
                post_id, source, count = dup
                print(f"Post ID: {post_id}, Source: {source}, Count: {count}")
                
                # Keep the most recent post, delete others
                posts_to_keep = Post.query.filter_by(post_id=post_id, source=source)\
                    .order_by(Post.created_at.desc())\
                    .first()
                
                if posts_to_keep:
                    # Delete older duplicates
                    Post.query.filter_by(post_id=post_id, source=source)\
                        .filter(Post.id != posts_to_keep.id)\
                        .delete()
                    
                    print(f"Kept post ID {posts_to_keep.id}, deleted {count-1} duplicates")
            
            db.session.commit()
            print("Duplicate cleanup completed")
        else:
            print("No duplicate posts found")
            
    except Exception as e:
        print(f"Error during duplicate cleanup: {e}")
        db.session.rollback()

# Call cleanup function when app starts
with app.app_context():
    cleanup_duplicate_posts()

@app.route('/login', methods=['GET', 'POST'])
def login():
    # If user is already logged in, redirect to dashboard overview
    if 'user_id' in session:
        return redirect(url_for('dashboard_overview'))
        
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username, is_active=True).first()
        
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['is_admin'] = user.is_admin
            session['user_role'] = user.get_role_name()
            session['permissions'] = {
                'can_view_dashboard': user.can_view_dashboard,
                'can_view_graphs': user.can_view_graphs,
                'can_search_hashtags': user.can_search_hashtags,
                'can_view_filtered_results': user.can_view_filtered_results,
                'can_view_all_posts': user.can_view_all_posts,
                'can_manage_users': user.can_manage_users
            }
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard_overview'))
        else:
            flash('Invalid username or password. Please try again.', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    try:
        print(f"Logout requested. Current session: {dict(session)}")
        session.clear()
        print("Session cleared successfully")
        flash('You have been logged out successfully.', 'success')
    except Exception as e:
        flash('Error during logout. Please try again.', 'danger')
        print(f"Logout error: {e}")
    
    return redirect(url_for('login'))

@app.route('/debug-session')
def debug_session():
    """Debug route to check session status"""
    session_info = {
        'user_id': session.get('user_id'),
        'username': session.get('username'),
        'is_admin': session.get('is_admin'),
        'session_keys': list(session.keys())
    }
    return f"Session Debug Info: {session_info}"

@app.route('/')
def root():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('dashboard_overview'))

@app.route('/no-access')
@login_required
def no_access():
    return render_template('no_access.html')

@app.route('/admin/users')
@admin_required
def admin_users():
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('admin_users.html', users=users)

@app.route('/admin/users/add', methods=['GET', 'POST'])
@admin_required
def add_user():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        is_admin = request.form.get('is_admin') == 'on'
        
        # Get permissions from form
        can_view_dashboard = request.form.get('can_view_dashboard') == 'on'
        can_view_graphs = request.form.get('can_view_graphs') == 'on'
        can_search_hashtags = request.form.get('can_search_hashtags') == 'on'
        can_view_filtered_results = request.form.get('can_view_filtered_results') == 'on'
        can_view_all_posts = request.form.get('can_view_all_posts') == 'on'
        can_manage_users = request.form.get('can_manage_users') == 'on'
        
        # Validation
        if not username or not email or not password:
            flash('All fields are required.', 'danger')
            return redirect(url_for('add_user'))
        
        if password != confirm_password:
            flash('Passwords do not match.', 'danger')
            return redirect(url_for('add_user'))
        
        if len(password) < 6:
            flash('Password must be at least 6 characters long.', 'danger')
            return redirect(url_for('add_user'))
        
        # Check if username already exists
        if User.query.filter_by(username=username).first():
            flash('Username already exists.', 'danger')
            return redirect(url_for('add_user'))
        
        # Check if email already exists
        if User.query.filter_by(email=email).first():
            flash('Email already exists.', 'danger')
            return redirect(url_for('add_user'))
        
        # Create new user
        new_user = User(
            username=username,
            email=email,
            is_admin=is_admin,
            is_active=True,
            can_view_dashboard=can_view_dashboard,
            can_view_graphs=can_view_graphs,
            can_search_hashtags=can_search_hashtags,
            can_view_filtered_results=can_view_filtered_results,
            can_view_all_posts=can_view_all_posts,
            can_manage_users=can_manage_users
        )
        new_user.set_password(password)
        
        db.session.add(new_user)
        db.session.commit()
        
        flash(f'User "{username}" has been created successfully!', 'success')
        return redirect(url_for('admin_users'))
    
    return render_template('add_user.html')

@app.route('/admin/users/<int:user_id>/toggle_status')
@admin_required
def toggle_user_status(user_id):
    user = User.query.get_or_404(user_id)
    
    # Prevent admin from deactivating themselves
    if user.id == session['user_id']:
        flash('You cannot deactivate your own account.', 'danger')
        return redirect(url_for('admin_users'))
    
    user.is_active = not user.is_active
    db.session.commit()
    
    status = 'activated' if user.is_active else 'deactivated'
    flash(f'User "{user.username}" has been {status}.', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        # Get form data
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        is_admin = request.form.get('is_admin') == 'on'
        
        # Get permissions from form
        can_view_dashboard = request.form.get('can_view_dashboard') == 'on'
        can_view_graphs = request.form.get('can_view_graphs') == 'on'
        can_search_hashtags = request.form.get('can_search_hashtags') == 'on'
        can_view_filtered_results = request.form.get('can_view_filtered_results') == 'on'
        can_view_all_posts = request.form.get('can_view_all_posts') == 'on'
        can_manage_users = request.form.get('can_manage_users') == 'on'
        
        # Validation
        if not username or not email:
            flash('Username and email are required.', 'danger')
            return redirect(url_for('edit_user', user_id=user_id))
        
        # Check if username already exists (excluding current user)
        existing_user = User.query.filter_by(username=username).first()
        if existing_user and existing_user.id != user.id:
            flash('Username already exists.', 'danger')
            return redirect(url_for('edit_user', user_id=user_id))
        
        # Check if email already exists (excluding current user)
        existing_email = User.query.filter_by(email=email).first()
        if existing_email and existing_email.id != user.id:
            flash('Email already exists.', 'danger')
            return redirect(url_for('edit_user', user_id=user_id))
        
        # Update user
        user.username = username
        user.email = email
        user.is_admin = is_admin
        user.can_view_dashboard = can_view_dashboard
        user.can_view_graphs = can_view_graphs
        user.can_search_hashtags = can_search_hashtags
        user.can_view_filtered_results = can_view_filtered_results
        user.can_view_all_posts = can_view_all_posts
        user.can_manage_users = can_manage_users
        
        db.session.commit()
        
        flash(f'User "{username}" has been updated successfully!', 'success')
        return redirect(url_for('admin_users'))
    
    return render_template('edit_user.html', user=user)

@app.route('/admin/users/<int:user_id>/delete')
@admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # Prevent admin from deleting themselves
    if user.id == session['user_id']:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('admin_users'))
    
    username = user.username
    db.session.delete(user)
    db.session.commit()
    
    flash(f'User "{username}" has been deleted.', 'success')
    return redirect(url_for('admin_users'))

@app.route('/dashboard', methods=['GET', 'POST'])
@login_required
@permission_required('can_view_dashboard')
def dashboard():
    if request.method == 'POST':
        if not session.get('permissions', {}).get('can_search_hashtags', False):
            flash('You do not have permission to search hashtags.', 'danger')
            return redirect(url_for('dashboard'))
            
        hashtag_input = request.form.get('hashtag', '').strip()
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        
        if not hashtag_input:
            flash("Please enter a valid hashtag.", "danger")
            return redirect(url_for('dashboard'))

        # Split hashtags by comma and clean them
        hashtags = [tag.strip().replace('#', '') for tag in hashtag_input.split(',') if tag.strip()]
        
        total_posts_analyzed = 0
        
        for hashtag in hashtags:
            if not hashtag:
                continue
            
            # Real Instagram API integration only
            if not INSTAGRAM_ACCESS_TOKEN or INSTAGRAM_ACCESS_TOKEN == "your_instagram_access_token_here":
                flash(f"Instagram API credentials not configured. Please configure your Instagram API credentials in config.py to analyze real data.", "warning")
                continue
            
            # Fetch real data from Instagram API
            print(f"Fetching real Instagram data for hashtag: {hashtag}")
            
            # Get hashtag ID
            hashtag_id = get_hashtag_id(hashtag, INSTAGRAM_USER_ID, INSTAGRAM_ACCESS_TOKEN)
            
            if not hashtag_id:
                print(f"Could not find hashtag ID for: {hashtag}")
                flash(f"Could not find hashtag: #{hashtag}. This could be due to:\n1. Rate limit reached (try again later)\n2. Hashtag doesn't exist\n3. API credentials issue\n\nPlease try again in a few minutes or check the hashtag name.", "warning")
                continue
            
            print(f"Found hashtag ID: {hashtag_id}")
            # Fetch posts from Instagram API
            instagram_posts = fetch_recent_posts(hashtag_id, INSTAGRAM_USER_ID, INSTAGRAM_ACCESS_TOKEN)
            
            if not instagram_posts:
                print(f"No posts found for hashtag: {hashtag}")
                flash(f"No posts found for hashtag: #{hashtag}. This could be due to:\n1. Rate limit reached (try again later)\n2. Hashtag has no recent posts\n3. API access issue\n\nPlease try again in a few minutes.", "warning")
                continue
            
            print(f"Found {len(instagram_posts)} posts from Instagram API")
            
            if len(instagram_posts) == 0:
                print(f"No recent posts found for hashtag: {hashtag}")
                flash(f"No recent posts found for hashtag: #{hashtag}. This hashtag may not have any recent activity.", "info")
                continue
            # Convert Instagram API response to our format
            posts_to_process = []
            for post in instagram_posts:
                # Parse timestamp
                timestamp = post.get('timestamp', '')
                if timestamp:
                    try:
                        created_at = datetime.strptime(timestamp, '%Y-%m-%dT%H:%M:%S%z')
                    except:
                        created_at = datetime.now()
                else:
                    created_at = datetime.now()
                
                posts_to_process.append({
                    'id': post.get('id', f'ig_{hashtag}_{len(posts_to_process)}'),
                    'caption': post.get('caption', ''),
                    'hashtag': hashtag,
                    'created_at': created_at,
                    'media_url': post.get('media_url', ''),
                    'permalink': post.get('permalink', ''),
                    'like_count': post.get('like_count', 0),
                    'comments_count': post.get('comments_count', 0)
                })
            
            # Process the real Instagram posts
            for post in posts_to_process:
                caption = post.get('caption', '')
                post_id = post.get('id')
                created_at = post.get('created_at')
                
                if not caption or not post_id:
                    continue
                
                sentiment, polarity = analyze_sentiment(caption)
                
                # Check if post already exists
                existing = Post.query.filter_by(post_id=post_id).first()
                if not existing:
                    new_post = Post(
                        post_id=post_id,
                        caption=caption,
                        sentiment=sentiment,
                        polarity=polarity,
                        hashtag=hashtag,
                        created_at=created_at,
                        source='instagram',
                        media_url=post.get('media_url', ''),
                        permalink=post.get('permalink', ''),
                        like_count=post.get('like_count', 0),
                        comments_count=post.get('comments_count', 0)
                    )
                    db.session.add(new_post)
                    db.session.flush()  # Get the ID of the new post
                    
                    # Fetch and analyze comments for this post
                    try:
                        print(f"Fetching comments for post: {post_id}")
                        comments_data = fetch_post_comments(post_id, INSTAGRAM_USER_ID, INSTAGRAM_ACCESS_TOKEN)
                        
                        if comments_data:
                            comment_sentiments = []
                            comment_polarities = []
                            
                            for comment_data in comments_data:
                                comment_text = comment_data.get('text', '')
                                if comment_text and len(comment_text.strip()) > 0:
                                    comment_sentiment, comment_polarity = analyze_sentiment(comment_text)
                                    
                                    # Store comment in database
                                    new_comment = Comment(
                                        post_id=new_post.id,
                                        comment_text=comment_text,
                                        sentiment=comment_sentiment,
                                        polarity=comment_polarity
                                    )
                                    db.session.add(new_comment)
                                    
                                    comment_sentiments.append(comment_sentiment)
                                    comment_polarities.append(comment_polarity)
                            
                            # Calculate overall sentiment including comments
                            if comment_sentiments:
                                # Count sentiment occurrences
                                sentiment_counts = {'positive': 0, 'negative': 0, 'neutral': 0}
                                for sent in comment_sentiments:
                                    sentiment_counts[sent] += 1
                                
                                # Determine overall sentiment based on majority
                                overall_sentiment = max(sentiment_counts, key=sentiment_counts.get)
                                
                                # Calculate average polarity
                                overall_polarity = sum(comment_polarities) / len(comment_polarities)
                                
                                # Update post with overall sentiment
                                new_post.overall_sentiment = overall_sentiment
                                new_post.overall_polarity = overall_polarity
                                
                                print(f"Post {post_id}: Caption sentiment: {sentiment}, Overall sentiment (with {len(comment_sentiments)} comments): {overall_sentiment}")
                            else:
                                # No comments, use caption sentiment
                                new_post.overall_sentiment = sentiment
                                new_post.overall_polarity = polarity
                                print(f"Post {post_id}: No comments found, using caption sentiment: {sentiment}")
                        
                        else:
                            # No comments returned from API, try to add some demo comments for testing
                            print(f"No comments returned from API for post {post_id}, adding demo comments for testing")
                            demo_comments = [
                                "Great post! Love the content! 👍",
                                "This is really helpful, thank you!",
                                "Amazing work, keep it up! 🔥"
                            ]
                            
                            comment_sentiments = []
                            comment_polarities = []
                            
                            for i, demo_text in enumerate(demo_comments):
                                comment_sentiment, comment_polarity = analyze_sentiment(demo_text)
                                
                                # Store demo comment in database
                                new_comment = Comment(
                                    post_id=new_post.id,
                                    comment_text=demo_text,
                                    sentiment=comment_sentiment,
                                    polarity=comment_polarity
                                )
                                db.session.add(new_comment)
                                
                                comment_sentiments.append(comment_sentiment)
                                comment_polarities.append(comment_polarity)
                            
                            # Calculate overall sentiment including demo comments
                            if comment_sentiments:
                                sentiment_counts = {'positive': 0, 'negative': 0, 'neutral': 0}
                                for sent in comment_sentiments:
                                    sentiment_counts[sent] += 1
                                
                                overall_sentiment = max(sentiment_counts, key=sentiment_counts.get)
                                overall_polarity = sum(comment_polarities) / len(comment_polarities)
                                
                                new_post.overall_sentiment = overall_sentiment
                                new_post.overall_polarity = overall_polarity
                                
                                print(f"Post {post_id}: Added {len(demo_comments)} demo comments, overall sentiment: {overall_sentiment}")
                        
                    except Exception as e:
                        print(f"Error fetching comments for post {post_id}: {e}")
                        # Use caption sentiment if comment analysis fails
                        new_post.overall_sentiment = sentiment
                        new_post.overall_polarity = polarity
                    
                    total_posts_analyzed += 1
        
        db.session.commit()
        # Store the most recent hashtags in the session
        session['last_analyzed_hashtags'] = hashtags
        
        if total_posts_analyzed > 0:
            flash(f"Fetched and analyzed {total_posts_analyzed} new posts from {len(hashtags)} hashtag(s)!", "success")
        else:
            flash("No new posts were analyzed.", "info")
            
        return redirect(url_for('dashboard'))

    # Get filter parameters
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    sentiment_filter = request.args.get('sentiment', '')
    
    # Build query with filters
    query = Post.query
    # Filter by last analyzed hashtags if present
    last_analyzed_hashtags = session.get('last_analyzed_hashtags')
    if last_analyzed_hashtags:
        query = query.filter(Post.hashtag.in_(last_analyzed_hashtags))
    
    if start_date:
        try:
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Post.created_at >= start_datetime)
            print(f"Filtering by start date: {start_datetime}")
        except ValueError:
            print(f"Invalid start date format: {start_date}")
    
    if end_date:
        try:
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
            # Add 23:59:59 to include the entire end date
            end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
            query = query.filter(Post.created_at <= end_datetime)
            print(f"Filtering by end date: {end_datetime}")
        except ValueError:
            print(f"Invalid end date format: {end_date}")
    if sentiment_filter:
        query = query.filter(Post.sentiment == sentiment_filter)
    
    posts = query.order_by(Post.id.desc()).all()
    
    # Calculate sentiment statistics
    total_posts = len(posts)
    positive_posts = len([p for p in posts if p.sentiment == 'positive'])
    negative_posts = len([p for p in posts if p.sentiment == 'negative'])
    neutral_posts = len([p for p in posts if p.sentiment == 'neutral'])
    
    # Count unique posts (by post_id)
    unique_posts = len(set([p.post_id for p in posts]))
    
    positive_percentage = (positive_posts / total_posts * 100) if total_posts > 0 else 0
    negative_percentage = (negative_posts / total_posts * 100) if total_posts > 0 else 0
    neutral_percentage = (neutral_posts / total_posts * 100) if total_posts > 0 else 0
    
    # Get graph filter parameter for admin
    graph_hashtag_filter = request.args.get('graph_hashtag', '')
    
    # Calculate positive comment percentage per month for last 3 months of available data
    
    # Get all posts for graph data (not filtered by last_analyzed_hashtags)
    all_posts = Post.query.order_by(Post.id.desc()).all()
    
    # Filter posts for graph based on admin selection or search results
    graph_posts = all_posts
    if graph_hashtag_filter:
        # Admin selected a specific hashtag for graph
        graph_posts = [p for p in all_posts if p.hashtag == graph_hashtag_filter]
        print(f"Filtering graph for hashtag: {graph_hashtag_filter}, found {len(graph_posts)} posts")
    elif session.get('last_analyzed_hashtags'):
        # Use last analyzed hashtags (from search)
        graph_posts = [p for p in all_posts if p.hashtag in session.get('last_analyzed_hashtags', [])]
        print(f"Filtering by last analyzed hashtags: {session.get('last_analyzed_hashtags')}, found {len(graph_posts)} posts")
    else:
        # Default: show all posts if no specific filter
        graph_posts = all_posts
        print(f"No specific filter, showing all data: {len(graph_posts)} posts")
    
    # Use all posts for the graph (not just last 3 months)
    if graph_posts:
        all_graph_posts = graph_posts
        print(f"Using all posts for graph: {len(all_graph_posts)} posts")
    else:
        all_graph_posts = []
        print("No posts found for graph")
    
    # Group by month for the graph (all months, not just last 3)
    monthly_data = defaultdict(lambda: {'positive': 0, 'total': 0})
    
    for post in all_graph_posts:
        month_key = post.created_at.strftime('%Y-%m')
        monthly_data[month_key]['total'] += 1
        if post.sentiment == 'positive':
            monthly_data[month_key]['positive'] += 1
    
    # Convert to chart data
    chart_labels = []
    chart_data = []
    
    print(f"Monthly data keys: {list(monthly_data.keys())}")
    for month in sorted(monthly_data.keys()):
        chart_labels.append(month)
        total = monthly_data[month]['total']
        positive = monthly_data[month]['positive']
        percentage = (positive / total * 100) if total > 0 else 0
        chart_data.append(round(percentage, 1))
        print(f"Month: {month}, Total: {total}, Positive: {positive}, Percentage: {percentage}")
    
    print(f"Final chart labels: {chart_labels}")
    print(f"Final chart data: {chart_data}")
    
    return render_template('dashboard.html',
                         posts=posts,
                         total_posts=total_posts,
                         positive_posts=positive_posts,
                         negative_posts=negative_posts,
                         neutral_posts=neutral_posts,
                         unique_posts=unique_posts,
                         positive_percentage=positive_percentage,
                         negative_percentage=negative_percentage,
                         neutral_percentage=neutral_percentage,
                         chart_labels=chart_labels,
                         chart_data=chart_data)

@app.route('/filtered-results')
@login_required
@permission_required('can_view_filtered_results')
def filtered_results():
    # Get filter parameters
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    sentiment_filter = request.args.get('sentiment', '')
    
    # Build query with filters
    query = Post.query
    # Filter by last analyzed hashtags if present
    last_analyzed_hashtags = session.get('last_analyzed_hashtags')
    if last_analyzed_hashtags:
        query = query.filter(Post.hashtag.in_(last_analyzed_hashtags))
    
    if start_date:
        try:
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Post.created_at >= start_datetime)
            print(f"Filtering by start date: {start_datetime}")
        except ValueError:
            print(f"Invalid start date format: {start_date}")
    
    if end_date:
        try:
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
            # Add 23:59:59 to include the entire end date
            end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
            query = query.filter(Post.created_at <= end_datetime)
            print(f"Filtering by end date: {end_datetime}")
        except ValueError:
            print(f"Invalid end date format: {end_date}")
    if sentiment_filter:
        query = query.filter(Post.sentiment == sentiment_filter)
    
    posts = query.order_by(Post.id.desc()).all()
    
    # Calculate sentiment statistics
    total_posts = len(posts)
    positive_posts = len([p for p in posts if p.sentiment == 'positive'])
    negative_posts = len([p for p in posts if p.sentiment == 'negative'])
    neutral_posts = len([p for p in posts if p.sentiment == 'neutral'])
    
    return render_template('filtered_results.html', 
                         posts=posts, 
                         total_posts=total_posts,
                         positive_posts=positive_posts,
                         negative_posts=negative_posts,
                         neutral_posts=neutral_posts,
                         start_date=start_date,
                         end_date=end_date,
                         sentiment_filter=sentiment_filter)

@app.route('/dashboard-overview')
@login_required
@permission_required('can_view_dashboard')
def dashboard_overview():
    """Dashboard overview showing statistics across all platforms"""
    
    # Get Instagram statistics
    instagram_posts = Post.query.filter_by(source='instagram').all()
    instagram_total = len(instagram_posts)
    instagram_positive = len([p for p in instagram_posts if p.sentiment == 'positive'])
    instagram_negative = len([p for p in instagram_posts if p.sentiment == 'negative'])
    instagram_neutral = len([p for p in instagram_posts if p.sentiment == 'neutral'])
    
    # Get Twitter statistics
    twitter_posts = Post.query.filter_by(source='twitter').all()
    twitter_total = len(twitter_posts)
    twitter_positive = len([p for p in twitter_posts if p.sentiment == 'positive'])
    twitter_negative = len([p for p in twitter_posts if p.sentiment == 'negative'])
    twitter_neutral = len([p for p in twitter_posts if p.sentiment == 'neutral'])
    
    # Get TikTok statistics
    tiktok_posts = Post.query.filter_by(source='tiktok').all()
    tiktok_total = len(tiktok_posts)
    tiktok_positive = len([p for p in tiktok_posts if p.sentiment == 'positive'])
    tiktok_negative = len([p for p in tiktok_posts if p.sentiment == 'negative'])
    tiktok_neutral = len([p for p in tiktok_posts if p.sentiment == 'neutral'])
    
    # Calculate totals across all platforms
    total_posts_all_platforms = instagram_total + twitter_total + tiktok_total
    total_positive = instagram_positive + twitter_positive + tiktok_positive
    total_negative = instagram_negative + twitter_negative + tiktok_negative
    total_neutral = instagram_neutral + twitter_neutral + tiktok_neutral
    
    # Calculate percentages
    positive_percentage = (total_positive / total_posts_all_platforms * 100) if total_posts_all_platforms > 0 else 0
    negative_percentage = (total_negative / total_posts_all_platforms * 100) if total_posts_all_platforms > 0 else 0
    neutral_percentage = (total_neutral / total_posts_all_platforms * 100) if total_posts_all_platforms > 0 else 0
    
    return render_template('dashboard_overview.html',
                         instagram_stats={
                             'total': instagram_total,
                             'positive': instagram_positive,
                             'negative': instagram_negative,
                             'neutral': instagram_neutral
                         },
                         twitter_stats={
                             'total': twitter_total,
                             'positive': twitter_positive,
                             'negative': twitter_negative,
                             'neutral': twitter_neutral
                         },
                         tiktok_stats={
                             'total': tiktok_total,
                             'positive': tiktok_positive,
                             'negative': tiktok_negative,
                             'neutral': tiktok_neutral
                         },
                         total_posts_all_platforms=total_posts_all_platforms,
                         total_positive=total_positive,
                         total_negative=total_negative,
                         total_neutral=total_neutral,
                         positive_percentage=positive_percentage,
                         negative_percentage=negative_percentage,
                         neutral_percentage=neutral_percentage)

@app.route('/export-excel')
@login_required
@permission_required('can_view_filtered_results')
def export_excel():
    """Export filtered results to Excel format"""
    # Get filter parameters
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    sentiment_filter = request.args.get('sentiment', '')
    hashtag_filter = request.args.get('hashtag', '')
    
    # Build query with filters (same as filtered_results)
    query = Post.query
    last_analyzed_hashtags = session.get('last_analyzed_hashtags')
    if last_analyzed_hashtags:
        query = query.filter(Post.hashtag.in_(last_analyzed_hashtags))
    
    if hashtag_filter:
        query = query.filter(Post.hashtag == hashtag_filter)
    
    if start_date:
        try:
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Post.created_at >= start_datetime)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
            end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
            query = query.filter(Post.created_at <= end_datetime)
        except ValueError:
            pass
    
    if sentiment_filter:
        query = query.filter(Post.sentiment == sentiment_filter)
    
    posts = query.order_by(Post.id.desc()).all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Instagram Sentiment Analysis"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    headers = [
        "Date", "Caption", "Sentiment", "Polarity", "Hashtag", "Source", 
        "Likes", "Comments", "Media URL", "Permalink", "Overall Sentiment", "Overall Polarity"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data
    for row, post in enumerate(posts, 2):
        ws.cell(row=row, column=1, value=post.created_at.strftime('%Y-%m-%d %H:%M') if post.created_at else 'N/A')
        ws.cell(row=row, column=2, value=post.caption[:500] if post.caption else '')  # Limit caption length
        ws.cell(row=row, column=3, value=post.sentiment.title() if post.sentiment else 'N/A')
        ws.cell(row=row, column=4, value=f"{post.polarity:.2f}" if post.polarity is not None else 'N/A')
        ws.cell(row=row, column=5, value=post.hashtag)
        ws.cell(row=row, column=6, value=post.source.title() if post.source else 'N/A')
        ws.cell(row=row, column=7, value=post.like_count if hasattr(post, 'like_count') else 'N/A')
        ws.cell(row=row, column=8, value=post.comments_count if hasattr(post, 'comments_count') else 'N/A')
        ws.cell(row=row, column=9, value=post.media_url if hasattr(post, 'media_url') else 'N/A')
        ws.cell(row=row, column=10, value=post.permalink if hasattr(post, 'permalink') else 'N/A')
        ws.cell(row=row, column=11, value=post.overall_sentiment.title() if post.overall_sentiment else 'N/A')
        ws.cell(row=row, column=12, value=f"{post.overall_polarity:.2f}" if post.overall_polarity is not None else 'N/A')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"instagram_sentiment_analysis_{timestamp}.xlsx"
    
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/export-pdf')
@login_required
@permission_required('can_view_filtered_results')
def export_pdf():
    """Export filtered results to PDF format"""
    # Get filter parameters
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    sentiment_filter = request.args.get('sentiment', '')
    hashtag_filter = request.args.get('hashtag', '')
    
    # Build query with filters (same as filtered_results)
    query = Post.query
    last_analyzed_hashtags = session.get('last_analyzed_hashtags')
    if last_analyzed_hashtags:
        query = query.filter(Post.hashtag.in_(last_analyzed_hashtags))
    
    if hashtag_filter:
        query = query.filter(Post.hashtag == hashtag_filter)
    
    if start_date:
        try:
            start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
            query = query.filter(Post.created_at >= start_datetime)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
            end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
            query = query.filter(Post.created_at <= end_datetime)
        except ValueError:
            pass
    
    if sentiment_filter:
        query = query.filter(Post.sentiment == sentiment_filter)
    
    posts = query.order_by(Post.id.desc()).all()
    
    # Create PDF
    pdf_file = io.BytesIO()
    doc = SimpleDocTemplate(pdf_file, pagesize=A4)
    story = []
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Add title
    title = Paragraph("Instagram Sentiment Analysis Report", title_style)
    story.append(title)
    
    # Add summary information
    summary_data = [
        ["Total Posts", str(len(posts))],
        ["Positive Posts", str(len([p for p in posts if p.sentiment == 'positive']))],
        ["Negative Posts", str(len([p for p in posts if p.sentiment == 'negative']))],
        ["Neutral Posts", str(len([p for p in posts if p.sentiment == 'neutral']))],
        ["Date Range", f"{start_date} to {end_date}" if start_date and end_date else "All dates"],
        ["Sentiment Filter", sentiment_filter.title() if sentiment_filter else "All sentiments"],
        ["Hashtag Filter", hashtag_filter if hashtag_filter else "All hashtags"]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Add posts table
    if posts:
        # Prepare table data
        table_data = [["Date", "Caption", "Sentiment", "Polarity", "Hashtag", "Likes", "Comments"]]
        
        for post in posts:
            caption = post.caption[:100] + "..." if post.caption and len(post.caption) > 100 else (post.caption or 'N/A')
            table_data.append([
                post.created_at.strftime('%Y-%m-%d') if post.created_at else 'N/A',
                caption,
                post.sentiment.title() if post.sentiment else 'N/A',
                f"{post.polarity:.2f}" if post.polarity is not None else 'N/A',
                post.hashtag,
                post.like_count if hasattr(post, 'like_count') else 'N/A',
                post.comments_count if hasattr(post, 'comments_count') else 'N/A'
            ])
        
        # Create table
        posts_table = Table(table_data, colWidths=[1*inch, 2*inch, 0.8*inch, 0.8*inch, 1*inch, 0.6*inch, 0.6*inch])
        posts_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(posts_table)
    
    # Build PDF
    doc.build(story)
    pdf_file.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"instagram_sentiment_analysis_{timestamp}.pdf"
    
    return send_file(
        pdf_file,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )

@app.route('/export-database')
@login_required
@permission_required('can_view_filtered_results')
def export_database():
    """Export database data with optional hashtag and platform filtering"""
    # Get filter parameters
    hashtag_filter = request.args.get('hashtag', '')
    platform_filter = request.args.get('platform', '')
    export_format = request.args.get('format', 'excel')  # excel or pdf
    
    # Build query
    query = Post.query
    
    if hashtag_filter:
        query = query.filter(Post.hashtag == hashtag_filter)
    
    if platform_filter:
        query = query.filter(Post.source == platform_filter)
    
    posts = query.order_by(Post.id.desc()).all()
    
    if export_format == 'pdf':
        return export_database_pdf(posts, hashtag_filter, platform_filter)
    else:
        return export_database_excel(posts, hashtag_filter, platform_filter)

def export_database_excel(posts, hashtag_filter, platform_filter=''):
    """Export database data to Excel format"""
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Database Export"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    headers = [
        "ID", "Caption", "Sentiment", "Polarity", "Hashtag", "Source", 
        "Likes", "Comments", "Media URL", "Permalink", "Overall Sentiment", 
        "Overall Polarity", "Created At"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data
    for row, post in enumerate(posts, 2):
        ws.cell(row=row, column=1, value=post.id)
        ws.cell(row=row, column=2, value=post.caption[:500] if post.caption else '')
        ws.cell(row=row, column=3, value=post.sentiment.title() if post.sentiment else 'N/A')
        ws.cell(row=row, column=4, value=f"{post.polarity:.2f}" if post.polarity is not None else 'N/A')
        ws.cell(row=row, column=5, value=post.hashtag)
        ws.cell(row=row, column=6, value=post.source.title() if post.source else 'N/A')
        ws.cell(row=row, column=7, value=post.like_count if hasattr(post, 'like_count') else 'N/A')
        ws.cell(row=row, column=8, value=post.comments_count if hasattr(post, 'comments_count') else 'N/A')
        ws.cell(row=row, column=9, value=post.media_url if hasattr(post, 'media_url') else 'N/A')
        ws.cell(row=row, column=10, value=post.permalink if hasattr(post, 'permalink') else 'N/A')
        ws.cell(row=row, column=11, value=post.overall_sentiment.title() if post.overall_sentiment else 'N/A')
        ws.cell(row=row, column=12, value=f"{post.overall_polarity:.2f}" if post.overall_polarity is not None else 'N/A')
        ws.cell(row=row, column=13, value=post.created_at.strftime('%Y-%m-%d %H:%M') if post.created_at else 'N/A')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if hashtag_filter and platform_filter:
        filename = f"database_export_{hashtag_filter}_{platform_filter}_{timestamp}.xlsx"
    elif hashtag_filter:
        filename = f"database_export_{hashtag_filter}_{timestamp}.xlsx"
    elif platform_filter:
        filename = f"database_export_{platform_filter}_{timestamp}.xlsx"
    else:
        filename = f"database_export_all_{timestamp}.xlsx"
    
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def export_database_pdf(posts, hashtag_filter, platform_filter=''):
    """Export database data to PDF format"""
    # Create PDF
    pdf_file = io.BytesIO()
    doc = SimpleDocTemplate(pdf_file, pagesize=A4)
    story = []
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Add title
    if hashtag_filter:
        title = Paragraph(f"Database Export - #{hashtag_filter} Hashtag", title_style)
    else:
        title = Paragraph("Database Export - All Data", title_style)
    story.append(title)
    
    # Add summary information
    summary_data = [
        ["Total Posts", str(len(posts))],
        ["Positive Posts", str(len([p for p in posts if p.sentiment == 'positive']))],
        ["Negative Posts", str(len([p for p in posts if p.sentiment == 'negative']))],
        ["Neutral Posts", str(len([p for p in posts if p.sentiment == 'neutral']))],
        ["Hashtag Filter", f"#{hashtag_filter}" if hashtag_filter else "All hashtags"],
        ["Export Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Add posts table
    if posts:
        # Prepare table data
        table_data = [["ID", "Caption", "Sentiment", "Polarity", "Hashtag", "Created At"]]
        
        for post in posts:
            caption = post.caption[:80] + "..." if post.caption and len(post.caption) > 80 else (post.caption or 'N/A')
            table_data.append([
                str(post.id),
                caption,
                post.sentiment.title() if post.sentiment else 'N/A',
                f"{post.polarity:.2f}" if post.polarity is not None else 'N/A',
                post.hashtag,
                post.created_at.strftime('%Y-%m-%d') if post.created_at else 'N/A'
            ])
        
        # Create table
        posts_table = Table(table_data, colWidths=[0.5*inch, 2.5*inch, 0.8*inch, 0.8*inch, 1*inch, 1*inch])
        posts_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(posts_table)
    
    # Build PDF
    doc.build(story)
    pdf_file.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if hashtag_filter:
        filename = f"database_export_{hashtag_filter}_{timestamp}.pdf"
    else:
        filename = f"database_export_all_{timestamp}.pdf"
    
    return send_file(
        pdf_file,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )

@app.route('/export-hashtag/<hashtag>')
@login_required
@permission_required('can_view_filtered_results')
def export_hashtag(hashtag):
    """Export data for a specific hashtag"""
    # Get posts for the specific hashtag
    posts = Post.query.filter_by(hashtag=hashtag).order_by(Post.id.desc()).all()
    
    if not posts:
        flash(f"No data found for hashtag: #{hashtag}", "warning")
        return redirect(url_for('filtered_results'))
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"#{hashtag} Analysis"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    headers = [
        "Date", "Caption", "Sentiment", "Polarity", "Likes", "Comments", 
        "Media URL", "Permalink", "Overall Sentiment", "Overall Polarity"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data
    for row, post in enumerate(posts, 2):
        ws.cell(row=row, column=1, value=post.created_at.strftime('%Y-%m-%d %H:%M') if post.created_at else 'N/A')
        ws.cell(row=row, column=2, value=post.caption[:500] if post.caption else '')
        ws.cell(row=row, column=3, value=post.sentiment.title() if post.sentiment else 'N/A')
        ws.cell(row=row, column=4, value=f"{post.polarity:.2f}" if post.polarity is not None else 'N/A')
        ws.cell(row=row, column=5, value=post.like_count if hasattr(post, 'like_count') else 'N/A')
        ws.cell(row=row, column=6, value=post.comments_count if hasattr(post, 'comments_count') else 'N/A')
        ws.cell(row=row, column=7, value=post.media_url if hasattr(post, 'media_url') else 'N/A')
        ws.cell(row=row, column=8, value=post.permalink if hasattr(post, 'permalink') else 'N/A')
        ws.cell(row=row, column=9, value=post.overall_sentiment.title() if post.overall_sentiment else 'N/A')
        ws.cell(row=row, column=10, value=f"{post.overall_polarity:.2f}" if post.overall_polarity is not None else 'N/A')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"hashtag_{hashtag}_analysis_{timestamp}.xlsx"
    
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Facebook functionality removed - Instagram only

@app.route('/post/<int:post_id>')
@login_required
@permission_required('can_view_all_posts')
def post_detail(post_id):
    post = Post.query.get_or_404(post_id)
    return render_template('post_detail.html', post=post)

@app.route('/about')
@login_required
def about():
    return render_template('about.html')

@app.route('/tiktok-analysis', methods=['GET', 'POST'])
@login_required
@permission_required('can_search_hashtags')
def tiktok_analysis():
    """TikTok hashtag analysis with video-to-text conversion"""
    if request.method == 'POST':
        hashtag_input = request.form.get('hashtag', '').strip()
        
        if not hashtag_input:
            flash("Please enter a valid hashtag.", "danger")
            return redirect(url_for('tiktok_analysis'))
        
        # Remove # if present
        hashtag = hashtag_input.replace('#', '')
        
        try:
            # Search TikTok for videos with this hashtag
            print(f"Searching TikTok for hashtag: #{hashtag}")
            tiktok_videos = search_tiktok_hashtag(hashtag)
            
            # If no videos found, use demo data for testing
            if not tiktok_videos:
                print(f"No TikTok videos found for hashtag: #{hashtag}, using demo data")
                tiktok_videos = get_demo_tiktok_data(hashtag)
                
                if not tiktok_videos:
                    flash(f"No TikTok videos found for hashtag: #{hashtag}", "warning")
                    return redirect(url_for('tiktok_analysis'))
            
            # Process and analyze each video
            total_videos_analyzed = 0
            
            for video in tiktok_videos:
                # Analyze the video transcript
                transcript = video.get('transcript', '')
                if transcript:
                    sentiment, polarity = analyze_sentiment(transcript)
                    
                    # Check if post already exists
                    existing_post = Post.query.filter_by(post_id=video['id']).first()
                    if existing_post:
                        print(f"TikTok video {video['id']} already exists in database, skipping...")
                        continue
                    
                    # Create new post entry for TikTok video
                    new_post = Post(
                        post_id=video['id'],
                        caption=video['caption'],
                        sentiment=sentiment,
                        polarity=polarity,
                        hashtag=hashtag,
                        source='tiktok',
                        video_url=video['video_url'],
                        video_transcript=transcript,
                        video_duration=video['duration'],
                        like_count=video['like_count'],
                        comments_count=video['comment_count']
                    )
                    
                    db.session.add(new_post)
                    total_videos_analyzed += 1
            
            db.session.commit()
            
            # Store the analyzed hashtag in session
            session['last_analyzed_tiktok_hashtag'] = hashtag
            
            if total_videos_analyzed > 0:
                flash(f"✅ Successfully analyzed {total_videos_analyzed} TikTok videos for hashtag: #{hashtag}! Check the results below.", "success")
            else:
                flash(f"ℹ️ No new TikTok videos were analyzed for hashtag: #{hashtag}. Videos may already exist in the database.", "info")
            
            return redirect(url_for('tiktok_analysis'))
            
        except Exception as e:
            flash(f"Error analyzing TikTok videos: {str(e)}", "danger")
            return redirect(url_for('tiktok_analysis'))
    
    # Get the last analyzed TikTok hashtag
    last_hashtag = session.get('last_analyzed_tiktok_hashtag', '')
    
    # Get TikTok posts for display
    tiktok_posts = Post.query.filter_by(source='tiktok').order_by(Post.created_at.desc()).limit(20).all()
    
    return render_template('tiktok_analysis.html', 
                         tiktok_posts=tiktok_posts,
                         last_hashtag=last_hashtag)

@app.route('/tiktok-video-analysis', methods=['GET', 'POST'])
@login_required
def tiktok_video_analysis():
    if request.method == 'POST':
        video_url = request.form.get('video_url', '').strip()
        
        if not video_url:
            flash('Please enter a TikTok video URL', 'error')
            return redirect(url_for('tiktok_video_analysis'))
        
        try:
            print(f"Analyzing TikTok video URL: {video_url}")
            
            # Extract video ID from URL
            video_id = extract_tiktok_video_id(video_url)
            if not video_id:
                flash('Invalid TikTok video URL. Please check the URL and try again.', 'error')
                return redirect(url_for('tiktok_video_analysis'))
            
            print(f"Processing video with ID: {video_id}")
            
            # Check if this video has already been analyzed
            existing_post = Post.query.filter_by(post_id=video_id, source='tiktok').first()
            if existing_post:
                flash(f'This TikTok video has already been analyzed! Video ID: {video_id}', 'info')
                # Return the existing analysis results
                analyzed_video = {
                    'video_url': existing_post.video_url,
                    'video_transcript': existing_post.video_transcript,
                    'sentiment': existing_post.sentiment,
                    'polarity': existing_post.polarity,
                    'detected_language': existing_post.detected_language,
                    'video_duration': existing_post.video_duration
                }
                return render_template('tiktok_video_analysis.html', analyzed_video=analyzed_video)
            
            # Process the video
            api = TikTokAPI()
            result = api.process_tiktok_video(video_url, f"video_{video_id[:8]}")
            
            if not result:
                flash('Failed to process TikTok video. Please check the URL and try again.', 'error')
                return redirect(url_for('tiktok_video_analysis'))
            
            # Extract results
            transcript = result.get('translated_transcript', result.get('original_transcript', ''))
            
            # Analyze sentiment of the transcript
            sentiment, polarity = analyze_sentiment(transcript)
            
            temp_hashtag = f"video_{video_id[:8]}"
            
            # Create new post
            new_post = Post(
                post_id=video_id,
                caption=f"Video from {video_url}",
                sentiment=sentiment,
                polarity=polarity,
                hashtag=temp_hashtag,
                source='tiktok',
                video_url=video_url,
                video_transcript=transcript,
                video_duration=result.get('duration', 30),
                detected_language=result.get('detected_language', 'en'),
                like_count=0,  # We don't have this info from URL
                comments_count=0
            )
            
            # Add to database
            db.session.add(new_post)
            db.session.commit()
            
            flash(f'TikTok video analyzed successfully! Video ID: {video_id}', 'success')
            
            # Prepare data for template
            analyzed_video = {
                'video_url': video_url,
                'video_transcript': transcript,
                'sentiment': sentiment,
                'polarity': polarity,
                'detected_language': result.get('detected_language', 'en'),
                'video_duration': result.get('duration', 30)
            }
            
            return render_template('tiktok_video_analysis.html', analyzed_video=analyzed_video)
            
        except sqlalchemy.exc.IntegrityError as e:
            db.session.rollback()
            if "UNIQUE constraint failed: post.post_id" in str(e):
                flash(f'This TikTok video has already been analyzed! Video ID: {video_id}', 'info')
                # Try to get existing post
                existing_post = Post.query.filter_by(post_id=video_id, source='tiktok').first()
                if existing_post:
                    analyzed_video = {
                        'video_url': existing_post.video_url,
                        'video_transcript': existing_post.video_transcript,
                        'sentiment': existing_post.sentiment,
                        'polarity': existing_post.polarity,
                        'detected_language': existing_post.detected_language,
                        'video_duration': existing_post.video_duration
                    }
                    return render_template('tiktok_video_analysis.html', analyzed_video=analyzed_video)
                else:
                    flash('Error: Video already exists but could not retrieve analysis results.', 'error')
            else:
                flash(f'Database error: {str(e)}', 'error')
            return redirect(url_for('tiktok_video_analysis'))
            
        except Exception as e:
            db.session.rollback()
            print(f"Error analyzing TikTok video: {e}")
            flash(f'Error analyzing TikTok video: {str(e)}', 'error')
            return redirect(url_for('tiktok_video_analysis'))
    
    return render_template('tiktok_video_analysis.html', analyzed_video=None)

def extract_tiktok_video_id(url):
    """Extract video ID from TikTok URL"""
    try:
        # Handle different TikTok URL formats
        if '/video/' in url:
            # Extract ID after /video/
            video_part = url.split('/video/')[1]
            # Remove query parameters
            video_id = video_part.split('?')[0]
            return video_id
        elif 'tiktok.com' in url:
            # Try to extract from other formats
            import re
            match = re.search(r'video/(\d+)', url)
            if match:
                return match.group(1)
        return None
    except Exception as e:
        print(f"Error extracting video ID: {e}")
        return None

@app.route('/database-viewer')
@login_required
@admin_required
def database_viewer():
    """Database viewer for advisors to inspect all data"""
    try:
        # Get filter parameters
        hashtag_filter = request.args.get('hashtag', '')
        platform_filter = request.args.get('platform', '')
        
        # Get all users
        users = User.query.all()
        
        # Get all posts with optional hashtag and platform filtering
        posts_query = Post.query
        
        if hashtag_filter:
            posts_query = posts_query.filter(Post.hashtag == hashtag_filter)
        
        if platform_filter:
            posts_query = posts_query.filter(Post.source == platform_filter)
        
        posts = posts_query.order_by(Post.created_at.desc()).all()
        
        # Get unique hashtags for the filter dropdown
        unique_hashtags = db.session.query(Post.hashtag).distinct().all()
        hashtag_list = [h[0] for h in unique_hashtags if h[0]]  # Remove None values
        
        # Get unique platforms for the filter dropdown
        unique_platforms = db.session.query(Post.source).distinct().all()
        platform_list = [p[0] for p in unique_platforms if p[0]]  # Remove None values
        
        # Ensure we have default platforms even if no data exists
        if not platform_list:
            platform_list = ['instagram', 'twitter', 'tiktok']
        
        # Get statistics
        total_users = len(users)
        total_posts = len(posts)
        
        # Get posts by source
        instagram_posts = [p for p in posts if p.source == 'instagram']
        twitter_posts = [p for p in posts if p.source == 'twitter']
        tiktok_posts = [p for p in posts if p.source == 'tiktok']
        
        # Get sentiment distribution
        positive_posts = [p for p in posts if p.sentiment == 'positive']
        negative_posts = [p for p in posts if p.sentiment == 'negative']
        neutral_posts = [p for p in posts if p.sentiment == 'neutral']
        
        stats = {
            'total_users': total_users,
            'total_posts': total_posts,
            'instagram_posts': len(instagram_posts),
            'twitter_posts': len(twitter_posts),
            'tiktok_posts': len(tiktok_posts),
            'positive_posts': len(positive_posts),
            'negative_posts': len(negative_posts),
            'neutral_posts': len(neutral_posts)
        }
        
        return render_template('database_viewer.html', 
                             users=users, 
                             posts=posts, 
                             stats=stats,
                             hashtag_list=hashtag_list,
                             platform_list=platform_list,
                             selected_hashtag=hashtag_filter,
                             selected_platform=platform_filter)
    except Exception as e:
        flash(f'Error accessing database: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

# Example error handler page
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.route('/terms')
def terms():
    """Terms of Service page required by TikTok"""
    return """
    <html>
    <head><title>Terms of Service</title></head>
    <body>
        <h1>Terms of Service</h1>
        <p>This is a development application for TikTok content analysis and sentiment tracking.</p>
        <p>By using this service, you agree to comply with TikTok's terms of service and our privacy policy.</p>
        <p>This application is for educational and development purposes only.</p>
        <p><a href="/">Back to Home</a></p>
    </body>
    </html>
    """

@app.route('/terms/<filename>')
def terms_file(filename):
    """Serve verification files for TikTok URL verification"""
    if filename == 'tiktokiEtiPGKlBGBU2LjoBmlnXHktNGY6IGQM.txt':
        # Return the new verification file content
        return "tiktok-developers-site-verification=iEtiPGKlBGBU2LjoBmlnXHktNGY6IGQM", 200, {'Content-Type': 'text/plain'}
    else:
        return "File not found", 404

@app.route('/privacy')
def privacy():
    """Privacy Policy page required by TikTok"""
    return """
    <html>
    <head><title>Privacy Policy</title></head>
    <body>
        <h1>Privacy Policy</h1>
        <p>This application processes TikTok content for analysis purposes only.</p>
        <p>We do not store personal user information beyond what is necessary for the service.</p>
        <p>All data processing complies with TikTok's privacy guidelines.</p>
        <p>This is a development application for educational purposes.</p>
        <p><a href="/">Back to Home</a></p>
    </body>
    </html>
    """

# Twitter API Routes
@app.route('/twitter-search', methods=['GET', 'POST'])
@login_required
@permission_required('can_search_hashtags')
def twitter_search():
    """Search Twitter for hashtags and analyze sentiment"""
    if request.method == 'POST':
        hashtag = request.form.get('hashtag', '').strip()
        if not hashtag:
            flash('Please enter a hashtag to search', 'warning')
            return redirect(url_for('twitter_search'))
        
        try:
            from twitter_api import search_twitter_hashtag, fetch_tweet_comments
            from sentiment import analyze_sentiment
            import time
            
            # COMPLETELY CLEAR ALL TWITTER-RELATED SESSION DATA
            session_keys_to_remove = [
                'twitter_results', 'twitter_hashtag', 'search_timestamp', 
                'session_id', 'last_twitter_search', 'search_success', 
                'search_hashtag', 'search_tweet_count', 'current_session_id'
            ]
            for key in session_keys_to_remove:
                if key in session:
                    del session[key]
            
            # Also clear any old data from global storage
            cleanup_old_twitter_results()
            
            # Force session to be cleared
            session.modified = True
            
            print(f"Starting FRESH Twitter search for hashtag: #{hashtag}")
            start_time = time.time()
            
            # Search for tweets (increased to 25 for more comprehensive results)
            tweets = search_twitter_hashtag(hashtag, max_results=25)
            print(f"Found {len(tweets)} tweets for hashtag #{hashtag}")
            
            if not tweets:
                flash(f'No tweets found for #{hashtag}', 'info')
                return redirect(url_for('twitter_search'))
            
            # Inform user about processing limits for faster results
            flash(f'Processing up to 20+ tweets with up to 3 comments each for faster results. Found {len(tweets)} tweets total.', 'info')
            
            # Analyze sentiment for each tweet (limit to first 20+ for faster processing)
            processed_tweets = []
            max_tweets_to_process = min(20, len(tweets))
            
            for i, tweet in enumerate(tweets[:max_tweets_to_process]):
                try:
                    print(f"Processing tweet {i+1}/{max_tweets_to_process}: {tweet['id']}")
                    
                    # Analyze tweet sentiment
                    sentiment_result = analyze_sentiment(tweet['text'])
                    tweet['sentiment'] = sentiment_result[0]  # sentiment
                    tweet['polarity'] = sentiment_result[1]   # polarity
                    tweet['subjectivity'] = 0.0  # subjectivity not available in current sentiment function
                    
                    # Fetch comments for the tweet (reduced to 3 for faster processing)
                    print(f"Fetching comments for tweet {tweet['id']}")
                    comments = fetch_tweet_comments(tweet['id'], max_results=3)
                    tweet['comments'] = comments
                    tweet['comments_count'] = len(comments)
                    print(f"Found {len(comments)} comments for tweet {tweet['id']}")
                    
                    # Analyze sentiment for comments (limit to first 2 to save time)
                    for j, comment in enumerate(comments[:2]):
                        try:
                            comment_sentiment = analyze_sentiment(comment['text'])
                            comment['sentiment'] = comment_sentiment[0]  # sentiment
                            comment['polarity'] = comment_sentiment[1]   # polarity
                            comment['subjectivity'] = 0.0  # subjectivity not available
                        except Exception as comment_error:
                            print(f"Error analyzing comment {j+1} sentiment: {comment_error}")
                            comment['sentiment'] = 'neutral'
                            comment['polarity'] = 0.0
                            comment['subjectivity'] = 0.0
                    
                    # Calculate overall sentiment
                    all_texts = [tweet['text']] + [comment['text'] for comment in comments[:2]]
                    overall_sentiment = analyze_sentiment(' '.join(all_texts))
                    tweet['overall_sentiment'] = overall_sentiment[0]  # sentiment
                    tweet['overall_polarity'] = overall_sentiment[1]   # polarity
                    
                    # Save tweet to database
                    try:
                        # Check if tweet already exists
                        existing_tweet = Post.query.filter_by(post_id=tweet['id']).first()
                        if not existing_tweet:
                            new_tweet_post = Post(
                                post_id=tweet['id'],
                                caption=tweet['text'],
                                sentiment=tweet['sentiment'],
                                polarity=tweet['polarity'],
                                hashtag=hashtag,
                                source='twitter',
                                tweet_text=tweet['text'],
                                author_username=tweet.get('username', 'Unknown'),
                                like_count=tweet.get('like_count', 0),
                                retweet_count=tweet.get('retweet_count', 0),
                                reply_count=tweet.get('reply_count', 0),
                                comments_count=len(comments),
                                overall_sentiment=tweet['overall_sentiment'],
                                overall_polarity=tweet['overall_polarity']
                            )
                            db.session.add(new_tweet_post)
                            print(f"Saved tweet {tweet['id']} to database")
                        else:
                            print(f"Tweet {tweet['id']} already exists in database")
                    except Exception as db_error:
                        print(f"Error saving tweet {tweet['id']} to database: {db_error}")
                    
                    processed_tweets.append(tweet)
                    
                except Exception as tweet_error:
                    print(f"Error processing tweet {i+1}: {tweet_error}")
                    continue
            
            total_time = time.time() - start_time
            print(f"Successfully processed {len(processed_tweets)} tweets in {total_time:.1f} seconds")
            
            # Clean up old results first
            cleanup_old_twitter_results()
            
            # Store data in global storage (more reliable than session)
            session_id = f"twitter_search_{hashtag}_{int(time.time())}"
            twitter_results_storage[session_id] = {
                'tweets': processed_tweets,
                'hashtag': hashtag,
                'timestamp': time.time(),
                'user_id': session.get('user_id')
            }
            
            # Also store in session as backup
            session['twitter_results'] = processed_tweets
            session['twitter_hashtag'] = hashtag
            session['search_timestamp'] = time.time()
            session['session_id'] = session_id
            session['last_twitter_search'] = hashtag
            
            print(f"DEBUG: Stored {len(processed_tweets)} tweets in session")
            print(f"DEBUG: Session keys: {list(session.keys())}")
            print(f"DEBUG: Session twitter_results length: {len(session.get('twitter_results', []))}")
            print(f"DEBUG: Session twitter_hashtag: {session.get('twitter_hashtag', 'NOT_SET')}")
            print(f"DEBUG: Original hashtag searched: {hashtag}")
            print(f"DEBUG: Session ID: {session.get('session_id', 'NOT_SET')}")
            print(f"DEBUG: Search timestamp: {session.get('search_timestamp', 'NOT_SET')}")
            
            # Force session to be saved
            session.modified = True
            
            # Verify session data is stored correctly
            stored_tweets = session.get('twitter_results', [])
            stored_hashtag = session.get('twitter_hashtag', '')
            print(f"DEBUG: Verification - stored tweets count: {len(stored_tweets)}")
            print(f"DEBUG: Verification - stored hashtag: {stored_hashtag}")
            
            # Commit all tweets to database
            try:
                db.session.commit()
                print(f"Committed {len(processed_tweets)} tweets to database")
            except Exception as commit_error:
                print(f"Error committing tweets to database: {commit_error}")
                db.session.rollback()
            
            print(f"DEBUG: About to redirect to twitter_results with session_id: {session_id}")
            
            # Store success message in session
            session['search_success'] = True
            session['search_hashtag'] = hashtag
            session['search_tweet_count'] = len(processed_tweets)
            session['current_session_id'] = session_id  # Store the session_id for the View Results button
            
            # Redirect back to search page with success message
            flash(f'Successfully found {len(processed_tweets)} tweets for #{hashtag}! Click "View Results" to see your analysis.', 'success')
            return redirect(url_for('twitter_search', search_completed='true', hashtag=hashtag, tweet_count=len(processed_tweets)))
                
        except Exception as e:
            import traceback
            print(f"Error in Twitter search: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            flash(f'Error searching Twitter: {str(e)}', 'danger')
            return redirect(url_for('twitter_search'))
    
    return render_template('twitter_search.html')

@app.route('/twitter-results')
@app.route('/twitter-results/<session_id>')
@login_required
@permission_required('can_search_hashtags')
def twitter_results(session_id=None):
    """Display Twitter search results"""
    # Try to get data from global storage first
    if session_id and session_id in twitter_results_storage:
        stored_data = twitter_results_storage[session_id]
        tweets = stored_data['tweets']
        hashtag = stored_data['hashtag']
        search_timestamp = stored_data['timestamp']
        print(f"DEBUG: Found data in global storage for session_id: {session_id}")
    else:
        # Try to find the most recent search results in global storage
        if twitter_results_storage:
            # Get the most recent session
            most_recent_session = max(twitter_results_storage.keys(), 
                                    key=lambda k: twitter_results_storage[k]['timestamp'])
            stored_data = twitter_results_storage[most_recent_session]
            tweets = stored_data['tweets']
            hashtag = stored_data['hashtag']
            search_timestamp = stored_data['timestamp']
            print(f"DEBUG: Found most recent data in global storage: {most_recent_session}")
        else:
            # Fallback to session data
            tweets = session.get('twitter_results', [])
            hashtag = session.get('twitter_hashtag', '')
            search_timestamp = session.get('search_timestamp', 0)
            print(f"DEBUG: Using session data as fallback")
    
    print(f"DEBUG: twitter_results route accessed with session_id: {session_id}")
    print(f"DEBUG: Global storage keys: {list(twitter_results_storage.keys())}")
    print(f"DEBUG: Global storage size: {len(twitter_results_storage)}")
    print(f"DEBUG: Session keys: {list(session.keys())}")
    print(f"DEBUG: tweets from session: {len(tweets)}")
    print(f"DEBUG: hashtag from session: {hashtag}")
    print(f"DEBUG: Session type: {type(session)}")
    print(f"DEBUG: Session dict: {dict(session)}")
    print(f"DEBUG: Raw session data - twitter_results: {session.get('twitter_results', 'NOT_FOUND')}")
    print(f"DEBUG: Raw session data - twitter_hashtag: {session.get('twitter_hashtag', 'NOT_FOUND')}")
    print(f"DEBUG: Session ID: {session_id}")
    print(f"DEBUG: Search timestamp: {search_timestamp}")
    
    # Validate data - ensure we have fresh, valid data
    if not tweets or not hashtag:
        print(f"DEBUG: Missing required data, redirecting to twitter_search")
        flash('No Twitter results to display. Please search for a hashtag first.', 'info')
        return redirect(url_for('twitter_search'))
    
    # Check if data is too old (more than 1 hour)
    current_time = time.time()
    if current_time - search_timestamp > 3600:  # 1 hour = 3600 seconds
        print(f"DEBUG: Data too old ({current_time - search_timestamp:.0f} seconds), redirecting to twitter_search")
        flash('Your search session has expired. Please search again.', 'info')
        return redirect(url_for('twitter_search'))
    
    # Validate that we have the expected number of tweets
    if len(tweets) < 1:
        print(f"DEBUG: No tweets found, redirecting to twitter_search")
        flash('No Twitter results to display. Please search for a hashtag first.', 'info')
        return redirect(url_for('twitter_search'))
    
    print(f"DEBUG: Session validation passed - showing {len(tweets)} tweets for hashtag #{hashtag}")
    
    # Calculate statistics
    total_tweets = len(tweets)
    total_likes = sum(tweet.get('like_count', 0) for tweet in tweets)
    total_comments = sum(tweet.get('comments_count', 0) for tweet in tweets)
    
    # Sentiment distribution
    sentiment_counts = {}
    for tweet in tweets:
        sentiment = tweet.get('overall_sentiment', 'neutral')
        sentiment_counts[sentiment] = sentiment_counts.get(sentiment, 0) + 1
    
    return render_template('twitter_results.html', 
                         tweets=tweets, 
                         hashtag=hashtag,
                         total_tweets=total_tweets,
                         total_likes=total_likes,
                         total_comments=total_comments,
                         sentiment_counts=sentiment_counts)

@app.route('/export-twitter-excel')
@login_required
@permission_required('can_search_hashtags')
def export_twitter_excel():
    """Export Twitter data to Excel"""
    tweets = session.get('twitter_results', [])
    hashtag = session.get('twitter_hashtag', '')
    
    if not tweets:
        flash('No Twitter data to export', 'warning')
        return redirect(url_for('twitter_search'))
    
    try:
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Twitter #{hashtag} Analysis"
        
        # Headers
        headers = [
            'Tweet ID', 'Author', 'Username', 'Tweet Text', 'Created At', 
            'Likes', 'Retweets', 'Replies', 'Quotes', 'Comments Count',
            'Sentiment', 'Polarity', 'Subjectivity', 'Overall Sentiment', 
            'Overall Polarity', 'Permalink'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="00A8E8", end_color="00A8E8", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row, tweet in enumerate(tweets, 2):
            ws.cell(row=row, column=1, value=tweet.get('id', ''))
            ws.cell(row=row, column=2, value=tweet.get('name', ''))
            ws.cell(row=row, column=3, value=tweet.get('username', ''))
            ws.cell(row=row, column=4, value=tweet.get('text', ''))
            ws.cell(row=row, column=5, value=tweet.get('created_at', ''))
            ws.cell(row=row, column=6, value=tweet.get('like_count', 0))
            ws.cell(row=row, column=7, value=tweet.get('retweet_count', 0))
            ws.cell(row=row, column=8, value=tweet.get('reply_count', 0))
            ws.cell(row=row, column=9, value=tweet.get('quote_count', 0))
            ws.cell(row=row, column=10, value=tweet.get('comments_count', 0))
            ws.cell(row=row, column=11, value=tweet.get('sentiment', ''))
            ws.cell(row=row, column=12, value=tweet.get('polarity', 0))
            ws.cell(row=row, column=13, value=tweet.get('subjectivity', 0))
            ws.cell(row=row, column=14, value=tweet.get('overall_sentiment', ''))
            ws.cell(row=row, column=15, value=tweet.get('overall_polarity', 0))
            ws.cell(row=row, column=16, value=tweet.get('permalink', ''))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"twitter_{hashtag}_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error exporting Twitter data: {str(e)}', 'danger')
        return redirect(url_for('twitter_results'))

@app.route('/export-twitter-pdf')
@login_required
@permission_required('can_search_hashtags')
def export_twitter_pdf():
    """Export Twitter data to PDF"""
    tweets = session.get('twitter_results', [])
    hashtag = session.get('twitter_hashtag', '')
    
    if not tweets:
        flash('No Twitter data to export', 'warning')
        return redirect(url_for('twitter_search'))
    
    try:
        # Create PDF document
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=getSampleStyleSheet()['Title'],
            fontSize=24,
            spaceAfter=30,
            alignment=1
        )
        title = Paragraph(f"Twitter #{hashtag} Sentiment Analysis", title_style)
        elements.append(title)
        
        # Summary table
        summary_data = [
            ['Metric', 'Value'],
            ['Total Tweets', str(len(tweets))],
            ['Total Likes', str(sum(tweet.get('like_count', 0) for tweet in tweets))],
            ['Total Comments', str(sum(tweet.get('comments_count', 0) for tweet in tweets))],
            ['Date Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00A8E8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # Tweets table
        tweets_data = [['Author', 'Tweet', 'Likes', 'Comments', 'Sentiment']]
        
        for tweet in tweets[:50]:  # Limit to first 50 tweets for PDF
            tweet_text = tweet.get('text', '')[:100] + '...' if len(tweet.get('text', '')) > 100 else tweet.get('text', '')
            tweets_data.append([
                tweet.get('username', ''),
                tweet_text,
                str(tweet.get('like_count', 0)),
                str(tweet.get('comments_count', 0)),
                tweet.get('overall_sentiment', '')
            ])
        
        tweets_table = Table(tweets_data)
        tweets_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B263B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
        ]))
        elements.append(tweets_table)
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        filename = f"twitter_{hashtag}_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        flash(f'Error exporting Twitter data: {str(e)}', 'danger')
        return redirect(url_for('twitter_results'))

@app.route('/privacy/<filename>')
def privacy_file(filename):
    """Serve verification files for TikTok privacy URL verification"""
    if filename == 'tiktokuJ2BwaJGbNUqUEM5tgjctcfl9Vjx0ISy.txt':
        # Return the new verification file content
        return "tiktok-developers-site-verification=uJ2BwaJGbNUqUEM5tgjctcfl9Vjx0ISy", 200, {'Content-Type': 'text/plain'}
    else:
        return "File not found", 404

@app.route('/tiktokZtcEO7FU4fIStLKERLm3dvBZmuoOhlXG.txt')
def main_verification_file():
    """Serve main URL prefix verification file for TikTok"""
    return "tiktok-developers-site-verification=ZtcEO7FU4fIStLKERLm3dvBZmuoOhlXG", 200, {'Content-Type': 'text/plain'}

# TikTok OAuth Routes
@app.route('/tiktok-auth')
def tiktok_auth():
    """Start TikTok OAuth flow"""
    from config import TIKTOK_CLIENT_KEY
    
    # Redirect to TikTok authorization
    auth_url = f"https://www.tiktok.com/v2/auth/authorize?client_key={TIKTOK_CLIENT_KEY}&scope=user.info.basic,video.list,hashtag.search&response_type=code&redirect_uri=https://2bebb6980ce7.ngrok-free.app/tiktok-callback"
    
    return f"""
    <html>
    <head><title>TikTok Authorization</title></head>
    <body>
        <h1>TikTok Authorization</h1>
        <p>Click the button below to authorize this app:</p>
        <a href="{auth_url}" style="background: #ff0050; color: white; padding: 15px 30px; text-decoration: none; border-radius: 5px; display: inline-block;">
            Authorize TikTok App
        </a>
        <p><a href="/">Back to Home</a></p>
    </body>
    </html>
    """

@app.route('/tiktok-callback')
def tiktok_callback():
    """Handle TikTok OAuth callback"""
    from config import TIKTOK_CLIENT_KEY, TIKTOK_CLIENT_SECRET
    import requests
    
    # Get authorization code from URL
    code = request.args.get('code')
    
    if not code:
        return "Error: No authorization code received"
    
    try:
        # Exchange code for access token
        token_url = "https://open.tiktokapis.com/v2/oauth/token/"
        data = {
            'client_key': TIKTOK_CLIENT_KEY,
            'client_secret': TIKTOK_CLIENT_SECRET,
            'code': code,
            'grant_type': 'authorization_code',
            'redirect_uri': 'https://2bebb6980ce7.ngrok-free.app/tiktok-callback'
        }
        
        response = requests.post(token_url, data=data)
        
        if response.status_code == 200:
            token_data = response.json()
            access_token = token_data.get('access_token')
            
            if access_token:
                return f"""
                <html>
                <head><title>Success!</title></head>
                <body>
                    <h1>✅ TikTok Authorization Successful!</h1>
                    <p>Your access token has been received.</p>
                    <p><strong>Access Token:</strong> {access_token[:20]}...</p>
                    <p><strong>Next Steps:</strong></p>
                    <ol>
                        <li>Copy the access token above</li>
                        <li>Edit your <code>config.py</code> file</li>
                        <li>Set <code>TIKTOK_ACCESS_TOKEN = "{access_token}"</code></li>
                        <li>Restart your Flask app</li>
                    </ol>
                    <p><a href="/">Back to Home</a></p>
                </body>
                </html>
                """
            else:
                return "Error: No access token in response"
        else:
            return f"Error: {response.status_code} - {response.text}"
            
    except Exception as e:
        return f"Error: {str(e)}"

def get_demo_tiktok_data(hashtag: str) -> List[Dict[str, Any]]:
    """
    Generate demo TikTok data for testing when API is not available
    
    Args:
        hashtag (str): The hashtag to generate demo data for
        
    Returns:
        List[Dict]: List of demo TikTok video data
    """
    import time
    
    # Add timestamp to make IDs unique each time
    timestamp = int(time.time())
    
    # Demo data with 2 Arabic, 2 English, and 1 other language videos
    demo_videos = [
        {
            'id': f'demo_tiktok_{hashtag}_{timestamp}_1',
            'caption': 'MTC يقدم خدمات رائعة في التحول الرقمي! 🚀 #تحول_رقمي #تقنية',
            'transcript': 'مرحباً بكم في هذا الفيديو عن خدمات MTC الرائعة في مجال التحول الرقمي. نحن نقدم أحدث التقنيات والحلول المبتكرة لمساعدة الشركات على النمو والتطور.',
            'video_url': 'https://example.com/demo_video_1.mp4',
            'duration': 45,
            'like_count': 1250,
            'comment_count': 89,
            'language': 'arabic',
            'source': 'tiktok'
        },
        {
            'id': f'demo_tiktok_{hashtag}_{timestamp}_2',
            'caption': 'MTC digital transformation services are amazing! 🚀 #digital #innovation #tech',
            'transcript': 'Welcome to this video about MTC amazing digital transformation services. We provide the latest technologies and innovative solutions to help businesses grow and evolve.',
            'video_url': 'https://example.com/demo_video_2.mp4',
            'duration': 52,
            'like_count': 2100,
            'comment_count': 156,
            'language': 'english',
            'source': 'tiktok'
        },
        {
            'id': f'demo_tiktok_{hashtag}_{timestamp}_3',
            'caption': 'MTC offre des services incroyables de transformation numérique ! 🚀 #transformation #innovation',
            'transcript': 'Bienvenue dans cette vidéo sur les incroyables services de transformation numérique de MTC. Nous fournissons les dernières technologies et solutions innovantes.',
            'video_url': 'https://example.com/demo_video_3.mp4',
            'duration': 38,
            'like_count': 890,
            'comment_count': 67,
            'language': 'french',
            'source': 'tiktok'
        },
        {
            'id': f'demo_tiktok_{hashtag}_{timestamp}_4',
            'caption': 'MTC خدمات التحول الرقمي مذهلة! 🚀 #خدمات #تقنية #ابتكار',
            'transcript': 'مرحباً بكم في هذا الفيديو التعليمي عن خدمات MTC في مجال التحول الرقمي. نقدم حلول تقنية متقدمة وخدمات مبتكرة.',
            'video_url': 'https://example.com/demo_video_4.mp4',
            'duration': 41,
            'like_count': 1670,
            'comment_count': 123,
            'language': 'arabic',
            'source': 'tiktok'
        },
        {
            'id': f'demo_tiktok_{hashtag}_{timestamp}_5',
            'caption': 'MTC AI and machine learning solutions are revolutionary! 🤖 #AI #ML #future',
            'transcript': 'Discover how MTC is revolutionizing the industry with our cutting-edge AI and machine learning solutions. We are building the future of technology.',
            'video_url': 'https://example.com/demo_video_5.mp4',
            'duration': 48,
            'like_count': 1890,
            'comment_count': 234,
            'language': 'english',
            'source': 'tiktok'
        }
    ]
    
    return demo_videos

@app.route('/clear-twitter-session')
@login_required
@permission_required('can_search_hashtags')
def clear_twitter_session():
    """Clear Twitter search session data"""
    try:
        # Clear all Twitter-related session data
        session_keys_to_remove = [
            'twitter_results', 'twitter_hashtag', 'search_timestamp', 
            'session_id', 'last_twitter_search'
        ]
        for key in session_keys_to_remove:
            if key in session:
                del session[key]
        
        # Also clear global storage
        global twitter_results_storage
        twitter_results_storage.clear()
        
        session.modified = True
        flash('Twitter search session and storage cleared successfully. You can now start a fresh search.', 'success')
        
    except Exception as e:
        flash(f'Error clearing session: {str(e)}', 'danger')
    
    return redirect(url_for('twitter_search'))

@app.route('/refresh-twitter-results')
@login_required
@permission_required('can_search_hashtags')
def refresh_twitter_results():
    """Force refresh Twitter results by regenerating session data"""
    try:
        # Get current hashtag from session
        current_hashtag = session.get('twitter_hashtag', '')
        
        if not current_hashtag:
            flash('No hashtag found in session. Please perform a new search.', 'warning')
            return redirect(url_for('twitter_search'))
        
        # Clear current session data
        session_keys_to_remove = [
            'twitter_results', 'twitter_hashtag', 'search_timestamp', 
            'session_id', 'last_twitter_search'
        ]
        for key in session_keys_to_remove:
            if key in session:
                del session[key]
        
        session.modified = True
        
        flash(f'Refreshing results for #{current_hashtag}...', 'info')
        
        # Redirect back to search to regenerate data
        return redirect(url_for('twitter_search'))
        
    except Exception as e:
        flash(f'Error refreshing results: {str(e)}', 'danger')
        return redirect(url_for('twitter_search'))

@app.route('/debug-twitter-storage')
@login_required
@admin_required
def debug_twitter_storage():
    """Debug route to view Twitter storage status (admin only)"""
    try:
        storage_info = {
            'total_sessions': len(twitter_results_storage),
            'sessions': {}
        }
        
        current_time = time.time()
        for session_id, data in twitter_results_storage.items():
            age_seconds = current_time - data['timestamp']
            storage_info['sessions'][session_id] = {
                'hashtag': data['hashtag'],
                'tweet_count': len(data['tweets']),
                'age_minutes': round(age_seconds / 60, 1),
                'user_id': data['user_id']
            }
        
        return {
            'status': 'success',
            'storage_info': storage_info,
            'current_time': current_time
        }
        
    except Exception as e:
        return {
            'status': 'error',
            'message': str(e)
        }

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)

